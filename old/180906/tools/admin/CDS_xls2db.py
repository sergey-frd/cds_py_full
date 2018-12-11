'''
This code uses the openpyxl package for playing around with excel using Python code
to convert complete excel workbook (all sheets) to an SQLite database
The code assumes that the first row of every sheet is the column name
Every sheet is stored in a separate table
The sheet name is assigned as the table name for every sheet
'''

import sqlite3
import openpyxl
from openpyxl import load_workbook
import re
import os, getopt, sys, StringIO
import fnmatch, string

########################################################################################
def del_old_file_V01(filepath):

    if os.path.isfile(filepath):
       try:
           os.remove(filepath)
           #print 'removed:', filepath
       except OSError:
           pass

#---------------------------------------------------
def slugify(text, lower=1):
    if lower == 1:
        text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text


#---------------------------------------------------

i=1
Input_File   = sys.argv[i]; i=i+1 
Out_File     = sys.argv[i]; i=i+1 
out_dir      = sys.argv[i]; i=i+1 
DB_Dictionary= sys.argv[i]; i=i+1

print "Input_File      =", Input_File
print "Out_File        =", Out_File  
print "out_dir         =", out_dir  
print "DB_Dictionary   =", DB_Dictionary  

exit

#---------------------------------------------------

del_old_file_V01(DB_Dictionary) 
f_DB_Dictionary  = open(DB_Dictionary, "w")

Dictionary = 'Dict_Table,Dict_Key'
#print "Dictionary=",Dictionary
f_DB_Dictionary.write(Dictionary+'\n')

#Replace with a database name
#con = sqlite3.connect('test.db')
#con = sqlite3.connect('Build_Base_Line_1.db')
con = sqlite3.connect(Out_File)
print "con=", con

#replace with the complete path to youe excel workbook
####wb = load_workbook(filename=r'Build_Base_Line_1.xlsx')
wb = load_workbook(filename=Input_File)
#wb = load_workbook(filename=r'CPU_Tests_Cases_Delta.xlsx')

#print "wb=", wb
#---------------------------------------------------
sheets = wb.get_sheet_names()
print "sheets=", sheets

for sheet in sheets:

    print "******************************************"
    print "sheet=",sheet, '"'+ sheet + '"'
    print "******************************************"

    OUT_CSV_FILE = os.path.join(out_dir,sheet+".csv")
    OUT_CSV_FILE = os.path.normpath(OUT_CSV_FILE)
    #print "OUT_CSV_FILE   =", OUT_CSV_FILE  
    del_old_file_V01(OUT_CSV_FILE) 
    f_OUT_CSV_FILE  = open(OUT_CSV_FILE, "w")

    ws = wb[sheet] 

    columns= []
    out_line=""

    query_d = 'drop table if exists ' + str(slugify(sheet))
    con.execute(query_d)
    con.commit()

    #con.execute("""drop table if exists str(slugify(sheet)) + '(ID INTEGER PRIMARY KEY AUTOINCREMENT'""")
    query = 'CREATE TABLE ' + str(slugify(sheet)) + '(ID INTEGER PRIMARY KEY AUTOINCREMENT'
    Dictionary=""


    #print "ws.rows[0]=", ws.rows[0]

    for row in ws.rows[0]:
        #print "row=", row

        ### query += ', ' + slugify(row.value) + ' TEXT'
        ### columns.append(slugify(row.value))
        ### print "row.value 0 =",row.value, slugify(row.value)

        query += ', ' + row.value + ' TEXT'
        Dictionary = sheet + ',' + row.value
        #print "Dictionary=",Dictionary
        f_DB_Dictionary.write(Dictionary+'\n')

        columns.append(row.value)
        #out_line=out_line + row.value +","
        out_line=out_line + unicode(row.value).strip() +","
        #print "row.value 0 =",row.value, out_line

        #print "querye=",query

    ### print "out_line =",out_line
    f_OUT_CSV_FILE.write(out_line+'\n')

    query += ');'

    #print "Dictionary END=",Dictionary
    print "querye END =  ",query
    query_ll = query.split(',')
    #for qq in query_ll:
    #    print "qq=", qq

    con.execute(query)

    tup = []
    for i, rows in enumerate(ws):
        tuprow = []
        if i == 0:
            continue

        out_line=""
        for row in rows:
            #print "row= 1", row
            #print "unicode(row.value).strip()= ", unicode(row.value).strip()
            out_line=out_line + unicode(row.value).strip() +","
            #out_line=out_line + row.value +","
            tuprow.append(unicode(row.value).strip()) if unicode(row.value).strip() != 'None' else tuprow.append('')

        ### print "out_line =",out_line
        f_OUT_CSV_FILE.write(out_line+'\n')
            #print "tuprow=", tuprow

        tup.append(tuple(tuprow))
        #print "tuple(tuprow)=",tuple(tuprow)


    f_OUT_CSV_FILE.close()

    #print "tup=",tup
    #tup_ll = tup.split(',')
    #for tt in tup:
    #    print "tt=", tt




    insQuery1 = 'INSERT INTO ' + str(slugify(sheet)) + '('
    insQuery2 = ''
    for col in columns:
        #print col, '     #     '+  sheet+"."+col,   '     #     "'+ col + '",'  '     #     ["'+ col + '"],'
        insQuery1 += col + ', '
        insQuery2 += '?, '
    insQuery1 = insQuery1[:-2] + ') VALUES('
    insQuery2 = insQuery2[:-2] + ')'
    insQuery = insQuery1 + insQuery2

    #print "insQuery=", insQuery
    #print "tup     =", tup

    insQuery_ll = insQuery.split(',')
    #for ii in insQuery_ll:
    #    print "ii=", ii
    #
    ##print "tup     =", tup
    #for tt in tup:
    #    print "tt=", tt

    con.executemany(insQuery, tup)
    con.commit()

    #for row in con:
    #        print (row)

con.close()

f_DB_Dictionary.close()

print "END con.close()"
#---------------------------------------------------
